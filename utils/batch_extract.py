"""
V53 Layer 1: 130家估值模型 批量提取脚本 (优化版)
========================================
从估值模型 Excel 中批量提取关键参数
输出: data/assumption_db.json (JSON Lines)
"""

from __future__ import annotations
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Optional

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
logger = logging.getLogger("batch_extract")

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl not installed.")
    sys.exit(1)

try:
    import xlrd
    _HAS_XLRD = True
except ImportError:
    _HAS_XLRD = False
    logger.warning("xlrd not installed, .xls will be skipped")


INDUSTRY_KEYWORDS = {
    "新能源车": ["新能源车", "比亚迪", "特斯拉", "蔚来", "小鹏", "理想", "宁德", "汽车"],
    "半导体": ["半导体", "芯片", "中芯", "华虹", "紫光", "韦尔", "北方华创", "兆易", "卓胜微", "华润微"],
    "互联网平台": ["互联网", "腾讯", "阿里", "美团", "百度", "京东", "拼多多", "哔哩", "快手", "字节", "网易", "小米"],
    "医药": ["医药", "恒瑞", "迈瑞", "药明", "复星", "智飞", "长春高新", "片仔癀", "白云山"],
    "消费": ["消费", "茅台", "五粮液", "伊利", "海天", "格力", "美的", "海尔", "安踏", "李宁", "蒙牛", "飞鹤"],
    "金融": ["金融", "银行", "保险", "证券", "招商银行", "平安", "中信", "工商银行", "宁波银行"],
    "地产": ["地产", "万科", "保利", "碧桂园", "龙湖", "华润置地", "融创"],
    "通信": ["通信", "中兴", "华为", "移动", "电信", "联通"],
    "军工": ["军工", "航发", "中航", "航天", "中船"],
    "化工": ["化工", "万华化学", "巴斯夫", "石化"],
    "机械": ["机械", "三一", "中联", "徐工"],
    "食品饮料": ["食品", "饮料", "乳业", "白酒", "啤酒", "调味"],
    "电子": ["电子", "立讯", "歌尔", "京东方", "TCL"],
    "煤炭": ["煤矿", "煤炭", "中国神华", "陕西煤业"],
    "公用事业": ["电力", "水务", "燃气", "长江电力", "华能"],
    "传媒": ["传媒", "分众", "芒果", "光线"],
}


def guess_industry(company_name: str, dir_name: str = "") -> str:
    text = (company_name + " " + dir_name).lower()
    for industry, keywords in INDUSTRY_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                return industry
    return "其他"


def extract_company_name(filename: str) -> str:
    name = os.path.splitext(filename)[0]
    name = re.sub(r'^[A-Z0-9]+\+', '', name)
    for suffix in ['财务估值模型', '估值模型', '财务预测估值模型', '财务预测']:
        name = name.replace(suffix, '')
    return name.strip()


# --- Fast cell scanning ---
def scan_for_numbers(ws, labels: list[str], max_row=80, max_col=20, offset_range=(1, 3), 
                     value_range=(None, None)) -> Optional[float]:
    """快速扫描表：找到标签后读取附近数值"""
    lo, hi = value_range
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip().lower()
                for label in labels:
                    if label.lower() in val:
                        row_idx, col_idx = cell.row, cell.column
                        for dc in range(offset_range[0], offset_range[1] + 1):
                            n = ws.cell(row=row_idx, column=col_idx + dc)
                            if n and n.value is not None:
                                try:
                                    v = float(n.value)
                                    if (lo is None or v >= lo) and (hi is None or v <= hi):
                                        return v
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception
                        # Also try below
                        for dr in range(1, 3):
                            n = ws.cell(row=row_idx + dr, column=col_idx)
                            if n and n.value is not None:
                                try:
                                    v = float(n.value)
                                    if (lo is None or v >= lo) and (hi is None or v <= hi):
                                        return v
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception
    return None


def find_label_row(ws, label: str, max_row=50) -> Optional[int]:
    """找到标签所在行"""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=5, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and label.lower() in cell.value.strip().lower():
                return cell.row
    return None


def compute_revenue_cagr(ws, max_row=50) -> Optional[float]:
    """从利润表计算营收CAGR"""
    rev_row = find_label_row(ws, "营业总收入", max_row)
    if not rev_row:
        rev_row = find_label_row(ws, "营业收入", max_row)
    if not rev_row:
        return None
    
    values = []
    for col in range(3, 8):  # Columns C-G
        cell = ws.cell(row=rev_row, column=col)
        if cell and cell.value is not None:
            try:
                v = float(cell.value)
                if v > 0:
                    values.append(v)
            except Exception:
                pass  # Layer 5: bare except replaced with Exception
    
    if len(values) >= 3:
        if values[0] > 0 and values[-1] > 0:
            n_years = len(values) - 1
            return (values[-1] / values[0]) ** (1/n_years) - 1
    return None


def find_gross_margin_from_sheet(ws, max_row=60) -> Optional[float]:
    """从利润表计算毛利率"""
    rev_row = find_label_row(ws, "营业总收入", max_row)
    if not rev_row:
        rev_row = find_label_row(ws, "营业收入", max_row)
    cost_row = find_label_row(ws, "营业成本", max_row)
    
    if rev_row and cost_row:
        for col in range(3, 8):
            r = ws.cell(row=rev_row, column=col)
            c = ws.cell(row=cost_row, column=col)
            if r and c and r.value and c.value:
                try:
                    rv, cv = float(r.value), float(c.value)
                    if rv > 0:
                        return round((rv - cv) / rv, 4)
                except Exception:
                    pass  # Layer 5: bare except replaced with Exception
    return None


def extract_from_xlsx(path: Path) -> Optional[dict]:
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = wb.sheetnames
    except Exception as e:
        logger.warning(f"Open failed: {path.name}: {e}")
        return None

    result = {
        "file": path.name, "company": extract_company_name(path.name),
        "dir": path.parent.name,
        "wacc": None, "terminal_growth": None, "revenue_cagr_3y": None,
        "gross_margin": None, "target_pe": None, "beta": None,
        "roe": None, "target_pb": None,
        "source_sheets": [], "extraction_quality": "none",
    }

    # Priority: DCF > 估值 > Cover > first sheets
    dcf_sheets = [s for s in sheets if 'dcf' in s.lower()]
    if not dcf_sheets:
        dcf_sheets = [s for s in sheets if '估值' in s.lower()]
    if not dcf_sheets:
        dcf_sheets = [s for s in sheets if 'cover' in s.lower()]
    if not dcf_sheets:
        dcf_sheets = sheets[:2]

    # Extract from DCF/估値 sheets
    for sname in dcf_sheets[:2]:
        ws = wb[sname]
        result["source_sheets"].append(sname)
        
        if result["wacc"] is None:
            v = scan_for_numbers(ws, ["wacc", "加权平均资本成本(WACC)", "加权平均资本成本", "资本成本"],
                                 offset_range=(1, 4), value_range=(0.01, 0.30))
            if v: result["wacc"] = round(v, 4)

        if result["terminal_growth"] is None:
            v = scan_for_numbers(ws, ["永续增长率g", "永续增长率", "terminal growth", "Terminal Growth", "增长率g"],
                                 offset_range=(1, 4), value_range=(0.0, 0.10))
            if v: result["terminal_growth"] = round(v, 4)

        if result["beta"] is None:
            v = scan_for_numbers(ws, ["β系数", "β", "beta", "Beta", "贝塔"],
                                 offset_range=(1, 3), value_range=(0.2, 2.5))
            if v: result["beta"] = round(v, 2)

        if result["target_pe"] is None:
            v = scan_for_numbers(ws, ["目标PE", "target PE", "Target PE", "PE（倍）", "PE(TTM)"],
                                 offset_range=(1, 3), value_range=(3, 150))
            if v: result["target_pe"] = round(v, 1)

        if result["roe"] is None:
            v = scan_for_numbers(ws, ["roe", "ROE", "净资产收益率"],
                                 offset_range=(1, 3), value_range=(0.01, 0.50))
            if v: result["roe"] = round(v, 4)

    # Income statement sheets for CAGR and gross margin
    inc_sheets = [s for s in sheets if '利润' in s or 'income' in s.lower()]
    for sname in inc_sheets[:2]:
        ws = wb[sname]
        if result["revenue_cagr_3y"] is None:
            cagr = compute_revenue_cagr(ws)
            if cagr is not None:
                result["revenue_cagr_3y"] = round(cagr, 4)
                result["source_sheets"].append(sname)
        if result["gross_margin"] is None:
            gm = find_gross_margin_from_sheet(ws)
            if gm is not None:
                result["gross_margin"] = gm
                result["source_sheets"].append(sname)

    wb.close()

    found = sum(1 for k in ["wacc","terminal_growth","revenue_cagr_3y","gross_margin","target_pe","beta"] if result.get(k))
    if found >= 5: result["extraction_quality"] = "high"
    elif found >= 3: result["extraction_quality"] = "medium"
    elif found >= 1: result["extraction_quality"] = "low"

    result["industry"] = guess_industry(result["company"], result["dir"])
    return result


def extract_from_xls(path: Path) -> Optional[dict]:
    if not _HAS_XLRD:
        return None
    try:
        wb = xlrd.open_workbook(str(path))
        sheets = wb.sheet_names()
    except Exception:
        return None

    result = {
        "file": path.name, "company": extract_company_name(path.name),
        "dir": path.parent.name,
        "wacc": None, "terminal_growth": None, "revenue_cagr_3y": None,
        "gross_margin": None, "target_pe": None, "beta": None,
        "roe": None, "target_pb": None,
        "source_sheets": [], "extraction_quality": "none",
    }

    dcf_sheets = [s for s in sheets if 'dcf' in s.lower() or '估值' in s.lower()]
    if not dcf_sheets:
        dcf_sheets = sheets[:2]

    for sname in dcf_sheets[:2]:
        ws = wb.sheet_by_name(sname)
        result["source_sheets"].append(sname)
        for r in range(min(ws.nrows, 100)):
            for c in range(min(ws.ncols, 18)):
                cell = ws.cell(r, c)
                if cell.ctype != 1:  # Not text
                    continue
                val = str(cell.value).strip().lower()
                
                if result["wacc"] is None and any(k in val for k in ['wacc', '加权平均资本成本', '资本成本']):
                    for dc in range(1, 4):
                        if c+dc < ws.ncols:
                            n = ws.cell(r, c+dc)
                            if n.ctype in (2, 3):
                                try:
                                    v = float(n.value)
                                    if 0.01 <= v <= 0.30:
                                        result["wacc"] = round(v, 4)
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception
                    # also check cell below
                    if r+1 < ws.nrows:
                        n = ws.cell(r+1, c)
                        if n.ctype in (2, 3):
                            try:
                                v = float(n.value)
                                if 0.01 <= v <= 0.30:
                                    result["wacc"] = round(v, 4)
                            except Exception:
                                pass  # Layer 5: bare except replaced with Exception

                if result["terminal_growth"] is None and any(k in val for k in ['永续增长率', 'terminal growth', '增长率g']):
                    for dc in range(1, 4):
                        if c+dc < ws.ncols:
                            n = ws.cell(r, c+dc)
                            if n.ctype in (2, 3):
                                try:
                                    v = float(n.value)
                                    if 0.0 <= v <= 0.10:
                                        result["terminal_growth"] = round(v, 4)
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception

                if result["beta"] is None and any(k in val for k in ['β', 'beta', '贝塔']):
                    for dc in range(1, 4):
                        if c+dc < ws.ncols:
                            n = ws.cell(r, c+dc)
                            if n.ctype in (2, 3):
                                try:
                                    v = float(n.value)
                                    if 0.2 <= v <= 2.5:
                                        result["beta"] = round(v, 2)
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception

                if result["target_pe"] is None and any(k in val for k in ['目标pe', 'target pe', 'pe（倍）', 'pe(t', 'pe估值']):
                    for dc in range(1, 4):
                        if c+dc < ws.ncols:
                            n = ws.cell(r, c+dc)
                            if n.ctype in (2, 3):
                                try:
                                    v = float(n.value)
                                    if 3 <= v <= 150:
                                        result["target_pe"] = round(v, 1)
                                except Exception:
                                    pass  # Layer 5: bare except replaced with Exception

    result["industry"] = guess_industry(result["company"], result["dir"])
    found = sum(1 for k in ["wacc","terminal_growth","revenue_cagr_3y","gross_margin","target_pe","beta"] if result.get(k))
    if found >= 3: result["extraction_quality"] = "medium"
    elif found >= 1: result["extraction_quality"] = "low"
    return result


def main():
    data_dir = Path(r"D:\Claude\1hao-analyst-v51\data\130家估值模型")
    output_path = Path(r"D:\Claude\1hao-analyst-v51\data\assumption_db.json")
    
    all_results = []
    
    for root, dirs, files in os.walk(str(data_dir)):
        for f in sorted(files):
            path = Path(root) / f
            if f.endswith('.xlsx'):
                r = extract_from_xlsx(path)
                if r:
                    all_results.append(r)
            elif f.endswith('.xls') and _HAS_XLRD:
                r = extract_from_xls(path)
                if r:
                    all_results.append(r)

    with open(str(output_path), 'w', encoding='utf-8') as f:
        for r in all_results:
            f.write(json.dumps(r, ensure_ascii=False) + '\n')

    # Summary
    qc = {}
    ic = {}
    for r in all_results:
        q = r.get("extraction_quality", "none")
        qc[q] = qc.get(q, 0) + 1
        ind = r.get("industry", "其他")
        ic[ind] = ic.get(ind, 0) + 1

    logger.info(f"=== Layer 1 提取完成 ===")
    logger.info(f"总记录: {len(all_results)} → {output_path}")
    logger.info(f"提取质量: {qc}")
    logger.info(f"行业覆盖 ({len(ic)}): {dict(sorted(ic.items(), key=lambda x: -x[1]))}")

    for k in ["wacc","terminal_growth","revenue_cagr_3y","gross_margin","target_pe","beta"]:
        vals = [r[k] for r in all_results if r.get(k)]
        if vals:
            logger.info(f"  {k}: {len(vals)} values, mean={sum(vals)/len(vals):.4f}, range=[{min(vals):.4f}, {max(vals):.4f}]")


if __name__ == "__main__":
    main()
