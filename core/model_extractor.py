"""模型提取器 — 从财务估值Excel提取DCF参数

读取 assets/models/ 和 assets/valuation/ 下的估值模型,
提取WACC/Beta/无风险利率/永续增长率/风险溢价等DCF参数,
存为 data/valuation_params.json 供数据驱动注入使用。
"""

import openpyxl, glob, json, re, os
from pathlib import Path
from typing import Optional

_ROOT = Path(__file__).resolve().parent.parent
MODEL_DIRS = [
    _ROOT / "assets" / "models",
    _ROOT / "assets" / "valuation",
]
OUTPUT = _ROOT / "data" / "valuation_params.json"


def extract_company_name(filepath: str) -> str:
    """从文件名提取公司名"""
    stem = Path(filepath).stem
    # 处理 "002049+紫光国微+财务估值模型" → "紫光国微"
    parts = stem.split('+')
    if len(parts) >= 2:
        return parts[1] if len(parts) == 2 else parts[1]
    # 处理 "宁德时代财务预测估值模型" → "宁德时代"
    for suffix in ['财务预测估值模型', '财务估值模型', '估值模型', '估值']:
        if suffix in stem:
            return stem.replace(suffix, '').strip()
    return stem[:10]


def extract_model(filepath: str) -> dict:
    """从单个Excel提取DCF参数"""
    result = {
        "file": os.path.basename(filepath),
        "company": extract_company_name(filepath),
        "params": {}
    }
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        
        for sheet_name in wb.sheetnames:
            sl = sheet_name.lower()
            if 'wacc' not in sl and 'dcf' not in sl:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for i, cell in enumerate(row):
                    if cell is None: continue
                    cell_str = str(cell).lower()
                    # 识别参数类型
                    param_type = None
                    if re.search(r'wacc|加权平均资本成本|加权.*资本成本', cell_str):
                        param_type = 'wacc'
                    elif re.search(r'\bbeta\b|β', cell_str):
                        param_type = 'beta'
                    elif re.search(r'无风险利率|风险[利率]|rf\b|r_f', cell_str):
                        param_type = 'risk_free'
                    elif re.search(r'永续增长|terminal.*growth|长期增长', cell_str):
                        param_type = 'growth'
                    elif re.search(r'风险溢价|equity.*risk|erp', cell_str):
                        param_type = 'market_risk'
                    
                    if not param_type:
                        continue
                    # 找同一行后面的数值
                    for j in range(i+1, min(i+8, len(row))):
                        v = row[j]
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            # 清洗异常值(永续增长率正常范围-5%~10%)
                            if param_type == 'growth' and (v < -0.1 or v > 0.15):
                                continue
                            if param_type == 'wacc' and (v < 0.01 or v > 0.25):
                                continue
                            if param_type == 'beta' and (v < 0.5 or v > 2.5):
                                continue
                            if param_type == 'risk_free' and (v < 0.01 or v > 0.08):
                                continue
                            if param_type == 'market_risk' and (v < 0.02 or v > 0.12):
                                continue
                            result["params"][param_type] = round(float(v), 4)
                            break
        wb.close()
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_all() -> list:
    """提取全部模型"""
    all_files = []
    for d in MODEL_DIRS:
        if d.exists():
            all_files += glob.glob(str(d / "**" / "*.xlsx"), recursive=True)
    
    results = []
    for f in all_files:
        r = extract_model(f)
        if r.get("params"):
            results.append(r)
    return results


def save_json(results: list):
    """保存为JSON,按公司名索引"""
    by_company = {}
    for r in results:
        company = r["company"]
        if not company:
            continue
        params = r["params"]
        if not params:
            continue
        if company not in by_company:
            by_company[company] = {}
        by_company[company].update(params)
    
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(by_company, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    return by_company


def get_params(company: str) -> dict:
    """按公司名查参数(供注入使用)"""
    if not OUTPUT.exists():
        return {}
    data = json.loads(OUTPUT.read_text(encoding="utf-8"))
    return data.get(company, {})


if __name__ == "__main__":
    results = extract_all()
    data = save_json(results)
    print(f"提取: {len(results)} 个模型")
    print(f"公司数: {len(data)}")
    print(f"输出: {OUTPUT}")
    for company, params in list(data.items())[:10]:
        print(f"  {company}: {params}")
