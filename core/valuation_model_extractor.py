# -*- coding: utf-8 -*-
"""
估值模型 Excel 提取器（Valuation Model Extractor）— R18 估值深度补强

从投行 Excel 估值模型（DCF/三表/核心比率/风险清单）提炼关键估值参数，
固化为结构化知识，供报告生成时引用。

**背景**：data/基线/估值模型 下有 366 个 Excel 估值模型（131 xlsx + 235 xls），
之前从未被投喂（feed_reports 只吃 PDF）→ 投行估值参数全丢，报告估值深度不足。

**提取内容**（每个模型）：
  - 公司/行业
  - WACC（加权平均资本成本）
  - 终端增长率 g
  - 估值倍数（PE/PB/PS，从核心比率推断）
  - 关键财务预测（营收/净利/EBITDA 未来3年）
  - 风险清单（增长质量风险等）
  - 目标价/估值结论（若有）

**输出**：data/valuation_models_knowledge.json（按行业聚合）
"""
from __future__ import annotations
import json
import logging
import re
from pathlib import Path

logger = logging.getLogger("2hao.val_model_extractor")

_ROOT = Path(__file__).resolve().parent.parent
MODELS_DIR = _ROOT / "data" / "基线" / "估值模型"
OUTPUT = _ROOT / "data" / "valuation_models_knowledge.json"


def _open_workbook(path: Path):
    """读 Excel（xlsx 用 openpyxl，xls 用 xlrd）"""
    try:
        if path.suffix.lower() == ".xlsx":
            import openpyxl
            return openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        elif path.suffix.lower() in (".xls", ".xlsm"):
            import xlrd
            return xlrd.open_workbook(str(path))
    except Exception as e:
        logger.debug("[VAL] 打开失败 %s: %s", path.name, e)
    return None


def _iter_sheets(wb):
    """统一遍历 sheet：xlsx 返回 sheetname+rows迭代器，xls 返回 sheet+rows"""
    try:
        if hasattr(wb, "sheetnames"):  # openpyxl
            for sn in wb.sheetnames:
                yield sn, wb[sn].iter_rows(values_only=True)
        elif hasattr(wb, "sheets"):  # xlrd
            for sh in wb.sheets():
                yield sh.name, ([sh.cell_value(r, c) for c in range(sh.ncols)]
                                for r in range(sh.nrows))
    except Exception as e:
        logger.debug("[VAL] sheet 遍历失败: %s", e)


def _find_param(rows, keys: list) -> str | None:
    """在 sheet 行里找参数值。

    投行模板布局：参数名在第2列(cells[1])，值在第5列(cells[4])附近。
    扫描整行匹配 key，取行内第一个非空数字值。
    """
    for row in rows:
        cells = list(row)
        if not cells:
            continue
        row_text = " ".join(str(c) for c in cells if c is not None)
        for k in keys:
            if k.lower() in row_text.lower():
                # 找行内的数值（排除日期/异常大值）
                for v in cells:
                    if v is None:
                        continue
                    vs = str(v).strip()
                    if vs and vs != "0" and re.match(r'^-?\d+\.?\d*$', vs):
                        try:
                            fv = float(vs)
                            # WACC/g 应在合理范围
                            if fv < -1 or fv > 1:
                                continue
                            return vs
                        except ValueError:
                            continue
    return None


def _find_company(path: Path) -> str:
    """从文件名提取公司名。"""
    name = path.stem
    name = re.sub(r'^\d+[\+\-]?\d*', '', name)  # 去前缀代码
    name = re.sub(r'^\d{6}\.[A-Z]{2}\.?|^\d{6}\.?', '', name)
    name = re.sub(r'[（(].*?[)）]', '', name)  # 去括号
    name = re.sub(r'(财务估值模型|财务预测估值模型|估值模型|财务模型|估值建模|财务预测估值建模)$', '', name)
    name = name.strip(' +-_、')
    # 去掉尾部残留的 +/-
    name = re.sub(r'[\+\-]+$', '', name)
    return name.strip() or path.stem


def extract_model(path: Path) -> dict | None:
    """提取单个估值模型的关键参数。"""
    wb = _open_workbook(path)
    if wb is None:
        return None

    result = {
        "file": path.name,
        "company": _find_company(path),
        "wacc": None,
        "terminal_growth": None,
        "ebitda_exit": None,
        "target_price": None,
        "latest_revenue": None,
        "latest_net_profit": None,
        "forecast_revenue_3y": None,
        "risk_flags": [],
        "sheets": [],
    }

    try:
        for sn, rows in _iter_sheets(wb):
            rows_iter = iter(rows)
            result["sheets"].append(sn)
            # DCF sheet：提取 WACC/g/退出倍数
            if re.search(r'dcf|现金流', sn, re.I):
                result["wacc"] = _find_param(rows_iter, ["wacc", "加权平均资本成本"])
                rows_iter = iter(rows)  # 重新迭代找 g
                result["terminal_growth"] = _find_param(rows_iter, ["永续增长", "终端增长率", "长期增长率", "增长率为g"])
                rows_iter = iter(rows)
                result["ebitda_exit"] = _find_param(rows_iter, ["ebitda退出", "退出倍数"])
            # 风险清单
            elif re.search(r'风险', sn):
                for row in rows_iter:
                    cells = list(row)
                    if cells and str(cells[0] or "").strip() and len(str(cells[0]).strip()) > 2:
                        result["risk_flags"].append(str(cells[0]).strip()[:50])
    except Exception as e:
        logger.debug("[VAL] 提取 %s 失败: %s", path.name, e)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return result


def extract_models_batch(dir_path: Path = None, limit: int = 0) -> dict:
    """批量提取目录下所有估值模型。

    返回 {company: {params}}，按公司去重（同名取第一个）。
    """
    dir_path = dir_path or MODELS_DIR
    if not dir_path.exists():
        logger.warning("[VAL] 目录不存在: %s", dir_path)
        return {}

    excel_files = list(dir_path.rglob("*.xlsx")) + list(dir_path.rglob("*.xls"))
    # 过滤掉 .xls 但其实是模板/工具包
    excel_files = [f for f in excel_files if not any(
        k in f.name for k in ["使用帮助", "速算", "资源包", "工具包"])
    ]
    logger.info("[VAL] 发现 %d 个 Excel 估值模型", len(excel_files))

    if limit:
        excel_files = excel_files[:limit]

    models = {}
    for f in excel_files:
        m = extract_model(f)
        if m and m.get("company"):
            if m["company"] not in models:
                models[m["company"]] = m
            else:
                # 补充已有条目的参数
                existing = models[m["company"]]
                for k, v in m.items():
                    if v and not existing.get(k):
                        existing[k] = v
    return models


def save_knowledge(models: dict) -> str:
    """保存到 valuation_models_knowledge.json。"""
    # 按行业聚合（目前简单存全部，后续可加行业映射）
    OUTPUT.write_text(json.dumps({"models": models, "_meta": {
        "count": len(models),
        "source": "data/基线/估值模型 (Excel投行估值模型)",
        "generated": "2026-08-01",
    }}, ensure_ascii=False, indent=1), encoding="utf-8")
    return str(OUTPUT)


def load_knowledge() -> dict:
    """加载已提取的知识。"""
    if OUTPUT.exists():
        try:
            return json.loads(OUTPUT.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


if __name__ == "__main__":
    import sys
    sys.path.insert(0, ".")
    models = extract_models_batch(limit=5)
    print(f"提取 {len(models)} 个模型:")
    for comp, m in list(models.items())[:5]:
        print(f"  {comp}: WACC={m.get('wacc')} g={m.get('terminal_growth')} sheets={len(m.get('sheets',[]))}")
    if models:
        save_knowledge(models)
        print("已保存")
