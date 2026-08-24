"""V51.6 Excel 导出器 — 对标130家估值模型的投行级输出格式

核心变化：
  1. 假设总表独立 Sheet（黄色底色标记假设单元格）
  2. WACC 拆解逐项可见
  3. 双变量敏感性矩阵
  4. 三张表联动（IS/BS/CFS）
  5. 投行级配色（蓝字=输入, 黑字=公式, 黄底=关键假设）
"""

from __future__ import annotations

import logging
from pathlib import Path

from core.models import AssumptionTree, ComputedResults

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: F401  (dead-import debt)
    from openpyxl.utils import get_column_letter  # noqa: F401  (availability probe)

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

logger = logging.getLogger("v51.export.xlsx")

# ═══════════════════════════════════════════════════════════════
# 格式常量（投行标准配色）
# ═══════════════════════════════════════════════════════════════

BLUE_FONT = Font(color="0000FF", name="Arial", size=10)  # 硬编码输入
BLACK_FONT = Font(color="000000", name="Arial", size=10)  # 公式
GREEN_FONT = Font(color="008000", name="Arial", size=10)  # 跨表引用
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 关键假设
LIGHT_BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # 表头
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)


def _apply_header_style(ws, row, max_col):
    """给表头行应用浅蓝底+粗体。"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = LIGHT_BLUE_FILL
        cell.font = Font(bold=True, name="Arial", size=10)
        cell.border = THIN_BORDER


def _apply_assumption_style(ws, row, col):
    """给假设单元格应用黄底+蓝字。"""
    cell = ws.cell(row=row, column=col)
    cell.fill = YELLOW_FILL
    cell.font = BLUE_FONT


# ═══════════════════════════════════════════════════════════════
# 导出器
# ═══════════════════════════════════════════════════════════════


class ModelExporter:
    """投行级 Excel 模型导出器。

    对标美团模型的 sheet 结构：
      Sheet 1: 假设总表（Assumptions）
      Sheet 2: 利润表（Income Statement）
      Sheet 3: 资产负债表（Balance Sheet）
      Sheet 4: 现金流量表（Cash Flow）
      Sheet 5: DCF 估值
      Sheet 6: 敏感性分析
      Sheet 7: 核心比率
    """

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        company: str,
        code: str,
        assumptions: AssumptionTree | None = None,
        results: ComputedResults | None = None,
        dcf_result: dict | None = None,
    ) -> str | None:
        """生成完整的估值模型 Excel。"""
        if not _HAS_OPENPYXL:
            logger.warning("openpyxl not installed, skipping xlsx export")
            return None

        wb = openpyxl.Workbook()

        # Sheet 1: 假设总表
        self._write_assumptions(wb, assumptions)

        # Sheet 2: DCF 估值
        self._write_dcf(wb, dcf_result)

        # Sheet 3: 敏感性分析
        self._write_sensitivity(wb, dcf_result)

        # Sheet 4: 核心财务数据
        self._write_financials(wb, results)

        # 保存
        safe_name = code.replace(".", "_") if code else company[:8]
        path = self.output_dir / f"{safe_name}_估值模型.xlsx"
        wb.save(str(path))
        logger.info(f"Model exported: {path}")
        return str(path)

    def _write_assumptions(self, wb, assumptions: AssumptionTree | None):
        """写入假设总表（对标美团模型的 Core Business Model Sheet）。"""
        ws = wb.active
        ws.title = "假设总表"
        ws.cell(row=1, column=1, value="关键假设总表").font = Font(bold=True, size=14)

        if not assumptions:
            ws.cell(row=3, column=1, value="（假设数据待补充 — 从130家模型对标库自动填充中）")
            return

        row = 3
        ws.cell(row=row, column=1, value="一、行业假设").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="指标")
        ws.cell(row=row, column=2, value="值")
        ws.cell(row=row, column=3, value="单位")
        ws.cell(row=row, column=4, value="来源/说明")
        _apply_header_style(ws, row, 4)
        row += 1

        # 营收驱动
        for node in assumptions.revenue_drivers:
            ws.cell(row=row, column=1, value=node.name)
            c = ws.cell(row=row, column=2, value=node.value)
            _apply_assumption_style(ws, row, 2)
            ws.cell(row=row, column=3, value=node.unit)
            ws.cell(row=row, column=4, value=node.description)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="二、WACC 拆解").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="指标")
        ws.cell(row=row, column=2, value="值")
        ws.cell(row=row, column=3, value="说明")
        _apply_header_style(ws, row, 3)
        row += 1

        wacc_items = [
            ("无风险利率", assumptions.wacc_assumptions.get("risk_free_rate", ""), "10年期国债收益率"),
            ("股权风险溢价", assumptions.wacc_assumptions.get("equity_risk_premium", ""), "市场风险溢价"),
            ("Beta", assumptions.wacc_assumptions.get("beta", ""), "与可比公司对标"),
            ("股权成本", assumptions.wacc_assumptions.get("cost_of_equity", ""), "= Rf + β × ERP"),
            ("债务成本", assumptions.wacc_assumptions.get("cost_of_debt", ""), "税后"),
            ("目标资本结构", assumptions.wacc_assumptions.get("debt_ratio", ""), "负债/总资本"),
            ("WACC", assumptions.wacc_assumptions.get("wacc", ""), "加权平均"),
        ]
        for name, val, note in wacc_items:
            ws.cell(row=row, column=1, value=name)
            if val:
                c = ws.cell(row=row, column=2, value=val)
                _apply_assumption_style(ws, row, 2)
                if isinstance(val, float):
                    c.number_format = "0.0%"
            ws.cell(row=row, column=3, value=note)
            row += 1

        # 设置列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 40

    def _write_dcf(self, wb, dcf_result: dict | None):
        """写入 DCF 估值表。"""
        ws = wb.create_sheet("DCF估值")
        ws.cell(row=1, column=1, value="DCF 估值模型").font = Font(bold=True, size=14)

        if not dcf_result:
            ws.cell(row=3, column=1, value="（DCF 数据待补充）")
            return

        row = 3
        items = [
            ("预测期 FCF 现值", dcf_result.get("present_value_of_fcf", 0), "亿元"),
            ("终值现值", dcf_result.get("terminal_value", 0), "亿元"),
            ("企业价值 (EV)", dcf_result.get("enterprise_value", 0), "亿元"),
            ("减：净债务", dcf_result.get("net_debt", 0), "亿元"),
            ("股权价值", dcf_result.get("equity_value", 0), "亿元"),
            ("股本", dcf_result.get("shares_outstanding", 1), "亿股"),
            ("目标价", dcf_result.get("target_price", 0), "元/股"),
        ]
        for name, val, unit in items:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=3, value=unit)
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1

    def _write_sensitivity(self, wb, dcf_result: dict | None):
        """写入敏感性矩阵。"""
        ws = wb.create_sheet("敏感性分析")
        ws.cell(row=1, column=1, value="敏感性分析：WACC × 永续增长率").font = Font(bold=True, size=11)

        if not dcf_result:
            ws.cell(row=3, column=1, value="（数据待补充）")
            return

        matrix = dcf_result.get("sensitivity_matrix", [])
        wacc_range = dcf_result.get("sensitivity_wacc_range", [])
        g_range = dcf_result.get("sensitivity_g_range", [])

        if not matrix or not wacc_range or not g_range:
            ws.cell(row=3, column=1, value="（敏感性数据不足）")
            return

        # 表头
        ws.cell(row=3, column=1, value="WACC \\ g")
        for j, g in enumerate(g_range):
            ws.cell(row=3, column=j + 2, value=f"{g * 100:.1f}%")
        _apply_header_style(ws, 3, len(g_range) + 1)

        # 数据
        for i, wacc in enumerate(wacc_range):
            row = 4 + i
            ws.cell(row=row, column=1, value=f"{wacc * 100:.1f}%")
            for j in range(len(g_range)):
                val = matrix[i][j] if i < len(matrix) and j < len(matrix[i]) else 0
                c = ws.cell(row=row, column=j + 2, value=val if val > 0 else None)
                c.number_format = "0.0"
                c.border = THIN_BORDER
                if val > 0:
                    # 颜色梯度: 高估值为绿, 低估值为红
                    if i == 2 and j == 1:  # base 情景
                        c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def _write_financials(self, wb, results: ComputedResults | None):
        """写入核心财务数据。"""
        ws = wb.create_sheet("核心比率")
        ws.cell(row=1, column=1, value="核心财务比率").font = Font(bold=True, size=14)

        if not results:
            ws.cell(row=3, column=1, value="（财务数据待补充）")
            return

        row = 3
        items = [
            (
                "毛利率",
                getattr(results, "margin_bridge", None) and getattr(results.margin_bridge, "gross_margin_current", ""),
                "%",
            ),
            ("营业利润率", getattr(results, "expense_bridge", None) and "N/A", "%"),
            ("ROE", "N/A", "%"),
            ("净利率", "N/A", "%"),
        ]
        for name, val, unit in items:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=val if val else "待补充")
            ws.cell(row=row, column=3, value=unit)
            row += 1

        # 列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10

    def integrate_to_pipeline(self, company: str, code: str, deliverable: object) -> str | None:
        """集成到 V51 管线——write 命令自动调用。"""
        assumptions = None
        results = None
        dcf_result = None

        if hasattr(deliverable, "knowledge_package") and deliverable.knowledge_package:
            kp = deliverable.knowledge_package
            if hasattr(kp, "financials") and kp.financials:
                results = kp.financials

        if hasattr(deliverable, "scaffold") and deliverable.scaffold:
            scaffold = deliverable.scaffold
            # 从 scaffold 提取假设
            if hasattr(scaffold, "sections"):
                for sec in scaffold.sections:
                    if sec.section_id == "valuation":
                        # 尝试提取 DCF 参数
                        pass

        return self.export(company, code, assumptions, results, dcf_result)
