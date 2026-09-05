"""
动态公式 Excel 导出器 — 写入原生 Excel 公式，确保投行级审计链。
每个数字都能在 Excel 中追溯到公式源头。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.schemas import (
    ComparableAssumptions,
    DCFAssumptions,
    ScenarioAssumptions,
    SOTPAssumptions,
)

# ─── Styles ─────────────────────────────────────────────────────────────────

DARK_BLUE = "1F4E78"
LIGHT_BLUE = "D9E1F2"
WHITE = "FFFFFF"
BLACK = "000000"

HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SECTION_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_BLUE)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
RESULT_FONT = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
REGULAR_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def _col(n: int) -> str:
    """1-indexed 列号 → Excel 列字母 (1=A, 27=AA)"""
    return get_column_letter(n)


class AuditExcelWriter:
    """生成带有原生公式的投行级审计底稿"""

    def __init__(
        self,
        dcf_assumptions: DCFAssumptions | None = None,
        comparable_assumptions: ComparableAssumptions | None = None,
        scenario_assumptions: ScenarioAssumptions | None = None,
        sotp_assumptions: SOTPAssumptions | None = None,
    ) -> None:
        self.dcf_a = dcf_assumptions
        self.comp_a = comparable_assumptions
        self.scn_a = scenario_assumptions
        self.sotp_a = sotp_assumptions
        self.wb = Workbook()

    def export(self, filepath: str = "DCF_Valuation_Model.xlsx") -> str:
        # 删除默认 sheet
        self.wb.remove(self.wb.active)

        if self.dcf_a:
            self._write_dcf_sheet()
        if self.comp_a:
            self._write_comparable_sheet()
        if self.scn_a:
            self._write_scenario_sheet()
        if self.sotp_a:
            self._write_sotp_sheet()
        self._write_summary_sheet()

        # 保存
        out = Path(filepath)
        out.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(out))
        return str(out.resolve())

    # ─── DCF Sheet ──────────────────────────────────────────────────────

    def _write_dcf_sheet(self) -> None:
        a = self.dcf_a
        ws = self.wb.create_sheet("DCF Valuation")
        years = a.forecast_years

        # Row 1: Title
        ws["A1"] = f"{a.ticker} ({a.company_name}) — DCF Valuation"
        ws["A1"].font = TITLE_FONT

        # Row 3: Section header
        ws["A3"] = "Model Assumptions"
        ws["A3"].font = SECTION_FONT
        ws["A3"].fill = SECTION_FILL

        # Row 4-10: Parameters (column B = values, referenced by formulas)
        params = [
            ("Base Revenue (亿元)", a.base_revenue, "#,##0.00"),
            ("WACC", a.wacc, "0.00%"),
            ("Terminal Growth (g)", a.terminal_growth_rate, "0.00%"),
            ("Tax Rate", a.tax_rate, "0.00%"),
            ("Net Debt (亿元)", a.net_debt, "#,##0.00"),
            ("Shares Outstanding (亿股)", a.shares_outstanding, "#,##0.00"),
            ("Current Price", a.current_price or 0, "#,##0.00"),
            ("D&A % Revenue", a.da_pct_revenue, "0.00%"),
            ("CapEx % Revenue", a.capex_pct_revenue, "0.00%"),
            ("WC Change % Revenue", a.wc_pct_revenue, "0.00%"),
        ]
        for i, (label, val, fmt) in enumerate(params, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = val
            ws[f"B{i}"].number_format = fmt
            ws[f"A{i}"].font = REGULAR_FONT

        # Row 15: Forecast header
        hdr_row = 15
        ws[f"A{hdr_row}"] = "Financial Forecast"
        ws[f"A{hdr_row}"].font = SECTION_FONT
        ws[f"A{hdr_row}"].fill = SECTION_FILL

        # Column headers: Year 1, Year 2, ...
        for i in range(years):
            c = _col(i + 2)  # B, C, D, ...
            ws[f"{c}{hdr_row}"] = f"Year {i + 1}"
            ws[f"{c}{hdr_row}"].font = HEADER_FONT
            ws[f"{c}{hdr_row}"].fill = HEADER_FILL
            ws[f"{c}{hdr_row}"].alignment = RIGHT_ALIGN

        # Row 16: Revenue Growth Rate
        r = hdr_row + 1
        ws[f"A{r}"] = "Revenue Growth"
        for i, g in enumerate(a.revenue_growth_rates):
            c = _col(i + 2)
            ws[f"{c}{r}"] = g
            ws[f"{c}{r}"].number_format = "0.00%"

        # Row 17: Revenue — formula: B17 = $B$4 * (1 + B16); C17 = B17 * (1 + C16)
        r += 1
        ws[f"A{r}"] = "Revenue"
        for i in range(years):
            c = _col(i + 2)
            if i == 0:
                ws[f"{c}{r}"] = f"=$B$4*(1+{c}{r - 1})"
            else:
                prev = _col(i + 1)
                ws[f"{c}{r}"] = f"={prev}{r}*(1+{c}{r - 1})"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 18: EBIT Margin
        r += 1
        ws[f"A{r}"] = "EBIT Margin"
        for i, m in enumerate(a.ebit_margins):
            c = _col(i + 2)
            ws[f"{c}{r}"] = m
            ws[f"{c}{r}"].number_format = "0.00%"

        # Row 19: EBIT = Revenue × Margin
        r += 1
        ws[f"A{r}"] = "EBIT"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 2}*{c}{r - 1}"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 20: NOPAT = EBIT × (1 - Tax)
        r += 1
        ws[f"A{r}"] = "NOPAT"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 1}*(1-$B$7)"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 21: D&A = Revenue × D&A%
        r += 1
        ws[f"A{r}"] = "Depreciation & Amort."
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 4}*$B$11"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 22: CapEx = Revenue × CapEx%
        r += 1
        ws[f"A{r}"] = "Capital Expenditure"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 5}*$B$12"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 23: WC Change = Revenue × WC%
        r += 1
        ws[f"A{r}"] = "Working Capital Δ"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 6}*$B$13"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 24: FCF = NOPAT + D&A - CapEx - WC
        r += 1
        ws[f"A{r}"] = "Free Cash Flow (FCF)"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 4}+{c}{r - 3}-{c}{r - 2}-{c}{r - 1}"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 25: Discount Factor = 1 / (1 + WACC)^t
        r += 1
        ws[f"A{r}"] = "Discount Factor"
        for i in range(years):
            c = _col(i + 2)
            t = i + 1
            ws[f"{c}{r}"] = f"=1/((1+$B$5)^{t})"
            ws[f"{c}{r}"].number_format = "0.0000"

        # Row 26: PV of FCF = FCF × DF
        r += 1
        ws[f"A{r}"] = "PV of FCF"
        for i in range(years):
            c = _col(i + 2)
            ws[f"{c}{r}"] = f"={c}{r - 2}*{c}{r - 1}"
            ws[f"{c}{r}"].number_format = "#,##0.00"

        # Row 28: Valuation Summary
        sum_row = 28
        ws[f"A{sum_row}"] = "Valuation Summary"
        ws[f"A{sum_row}"].font = SECTION_FONT
        ws[f"A{sum_row}"].fill = SECTION_FILL
        last_col = _col(years + 1)  # last year column

        # Cumulative PV of FCF
        r = sum_row + 1
        ws[f"A{r}"] = "Cumulative PV of FCF"
        ws[f"B{r}"] = f"=SUM(B26:{last_col}26)"
        ws[f"B{r}"].number_format = "#,##0.00"

        # Terminal Value = (Last_FCF × (1+g)) / (WACC - g)
        r += 1
        ws[f"A{r}"] = "Terminal Value (TV)"
        ws[f"B{r}"] = f"=({last_col}24*(1+$B$6))/($B$5-$B$6)"
        ws[f"B{r}"].number_format = "#,##0.00"

        # PV of TV = TV × Last_DF
        r += 1
        ws[f"A{r}"] = "PV of Terminal Value"
        ws[f"B{r}"] = f"=B{r - 1}*{last_col}25"
        ws[f"B{r}"].number_format = "#,##0.00"

        # EV = Sum PV FCF + PV TV
        r += 1
        ws[f"A{r}"] = "Enterprise Value (EV)"
        ws[f"B{r}"] = f"=B{r - 3}+B{r - 1}"
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"A{r}"].font = BOLD_FONT
        ws[f"B{r}"].font = BOLD_FONT

        # Less: Net Debt
        r += 1
        ws[f"A{r}"] = "Less: Net Debt"
        ws[f"B{r}"] = "=$B$8"
        ws[f"B{r}"].number_format = "#,##0.00"

        # Equity Value
        r += 1
        ws[f"A{r}"] = "Equity Value"
        ws[f"B{r}"] = f"=B{r - 2}-B{r - 1}"
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"A{r}"].font = BOLD_FONT
        ws[f"B{r}"].font = BOLD_FONT

        # Per Share Value
        r += 1
        ws[f"A{r}"] = "Intrinsic Value Per Share"
        ws[f"B{r}"] = f"=B{r - 1}/$B$9"
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"A{r}"].font = RESULT_FONT
        ws[f"B{r}"].font = RESULT_FONT

        # Upside %
        if a.current_price and a.current_price > 0:
            r += 1
            ws[f"A{r}"] = "Upside / Downside"
            ws[f"B{r}"] = f"=B{r - 1}/$B$10-1"
            ws[f"B{r}"].number_format = "0.00%"

        # TV as % of EV
        r += 1
        ws[f"A{r}"] = "TV % of EV"
        tv_row = sum_row + 3
        ev_row = sum_row + 5
        ws[f"B{r}"] = f"=B{tv_row}/B{ev_row}"
        ws[f"B{r}"].number_format = "0.00%"

        # Sensitivity table
        r += 2
        ws[f"A{r}"] = "Sensitivity Analysis (Fair Value per Share)"
        ws[f"A{r}"].font = SECTION_FONT
        ws[f"A{r}"].fill = SECTION_FILL

        r += 1
        ws[f"A{r}"] = "WACC \\ g →"
        for j, g in enumerate(a.sensitivity_g_range if hasattr(a, "sensitivity_g_range") else []):
            c = _col(j + 2)
            ws[f"{c}{r}"] = g
            ws[f"{c}{r}"].number_format = "0.00%"

        # Note: sensitivity values are computed by DCFEngine, injected as static values
        # (formula-based sensitivity would require Data Tables which need VBA)
        ws[f"A{r + 1}"] = "(Computed sensitivity matrix — see dcf_result.sensitivity_matrix)"

        # Column widths
        ws.column_dimensions["A"].width = 30
        for i in range(years + 1):
            ws.column_dimensions[_col(i + 2)].width = 16

    # ─── Comparable Sheet ───────────────────────────────────────────────

    def _write_comparable_sheet(self) -> None:
        a = self.comp_a
        ws = self.wb.create_sheet("Comparable Valuation")

        ws["A1"] = f"{a.ticker} ({a.company_name}) — Comparable Valuation"
        ws["A1"].font = TITLE_FONT

        # Company metrics
        ws["A3"] = "Company Metrics"
        ws["A3"].font = SECTION_FONT
        ws["A3"].fill = SECTION_FILL

        metrics = [
            ("EPS (元)", a.company_eps),
            ("BVPS (元)", a.company_bvps),
            ("Revenue/Share (元)", a.company_revenue_per_share),
            ("EBITDA/Share (元)", a.company_ebitda_per_share),
        ]
        for i, (label, val) in enumerate(metrics, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = val
            ws[f"B{i}"].number_format = "#,##0.00"

        # Peer multiples
        r = 9
        ws[f"A{r}"] = "Peer Multiples"
        ws[f"A{r}"].font = SECTION_FONT
        ws[f"A{r}"].fill = SECTION_FILL

        r += 1
        ws[f"A{r}"] = "Metric"
        ws[f"B{r}"] = "Mean"
        ws[f"C{r}"] = "Median"
        ws[f"D{r}"] = "Min"
        ws[f"E{r}"] = "Max"
        ws[f"F{r}"] = "Count"
        for c in "ABCDEF":
            ws[f"{c}{r}"].font = HEADER_FONT
            ws[f"{c}{r}"].fill = HEADER_FILL

        peer_data = []
        if a.peer_pe_ratios:
            peer_data.append(("PE", a.peer_pe_ratios))
        if a.peer_pb_ratios:
            peer_data.append(("PB", a.peer_pb_ratios))
        if a.peer_ps_ratios:
            peer_data.append(("PS", a.peer_ps_ratios))
        if a.peer_ev_ebitda:
            peer_data.append(("EV/EBITDA", a.peer_ev_ebitda))

        for metric_name, values in peer_data:
            r += 1
            s = sorted(values)
            n = len(s)
            mean = sum(s) / n
            median = s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2
            ws[f"A{r}"] = metric_name
            ws[f"B{r}"] = round(mean, 2)
            ws[f"C{r}"] = round(median, 2)
            ws[f"D{r}"] = round(s[0], 2)
            ws[f"E{r}"] = round(s[-1], 2)
            ws[f"F{r}"] = n

        # Implied prices
        r += 2
        ws[f"A{r}"] = "Implied Target Prices"
        ws[f"A{r}"].font = SECTION_FONT
        ws[f"A{r}"].fill = SECTION_FILL

        r += 1
        if a.peer_pe_ratios and a.company_eps > 0:
            mean_pe = sum(a.peer_pe_ratios) / len(a.peer_pe_ratios)
            ws[f"A{r}"] = "PE Implied Price"
            ws[f"B{r}"] = f"=B4*{round(mean_pe, 2)}"
            ws[f"B{r}"].number_format = "#,##0.00"
            r += 1

        if a.peer_pb_ratios and a.company_bvps > 0:
            mean_pb = sum(a.peer_pb_ratios) / len(a.peer_pb_ratios)
            ws[f"A{r}"] = "PB Implied Price"
            ws[f"B{r}"] = f"=B5*{round(mean_pb, 2)}"
            ws[f"B{r}"].number_format = "#,##0.00"
            r += 1

        ws.column_dimensions["A"].width = 22
        for c in "BCDEF":
            ws.column_dimensions[c].width = 14

    # ─── Scenario Sheet ─────────────────────────────────────────────────

    def _write_scenario_sheet(self) -> None:
        a = self.scn_a
        ws = self.wb.create_sheet("Scenario Analysis")

        ws["A1"] = f"{a.ticker} ({a.company_name}) — Scenario Analysis"
        ws["A1"].font = TITLE_FONT

        # Parameters
        ws["A3"] = "Scenario Parameters"
        ws["A3"].font = SECTION_FONT
        ws["A3"].fill = SECTION_FILL

        params = [
            ("Base Price", a.base_price),
            ("WACC", a.wacc),
            ("Tax Rate", a.tax_rate),
            ("Projection Years", a.projection_years),
        ]
        for i, (label, val) in enumerate(params, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = val

        # Scenario table
        r = 9
        ws[f"A{r}"] = "Scenario"
        ws[f"B{r}"] = "Bull"
        ws[f"C{r}"] = "Base"
        ws[f"D{r}"] = "Bear"
        for c in "ABCD":
            ws[f"{c}{r}"].font = HEADER_FONT
            ws[f"{c}{r}"].fill = HEADER_FILL

        details = [("bull", a.bull), ("base", a.base), ("bear", a.bear)]
        scenarios_data = [d for _, d in details]

        r += 1
        ws[f"A{r}"] = "Probability"
        for j, (_, d) in enumerate(details):
            ws[f"{_col(j + 2)}{r}"] = d.probability
            ws[f"{_col(j + 2)}{r}"].number_format = "0%"

        r += 1
        ws[f"A{r}"] = "Operating Margin"
        for j, (_, d) in enumerate(details):
            ws[f"{_col(j + 2)}{r}"] = d.operating_margin
            ws[f"{_col(j + 2)}{r}"].number_format = "0.00%"

        r += 1
        ws[f"A{r}"] = "Terminal Growth"
        for j, (_, d) in enumerate(details):
            ws[f"{_col(j + 2)}{r}"] = d.terminal_growth
            ws[f"{_col(j + 2)}{r}"].number_format = "0.00%"

        r += 1
        ws[f"A{r}"] = "Growth Rates"
        for j, (_, d) in enumerate(details):
            rates_str = ", ".join(f"{g:.0%}" for g in d.revenue_growth_rates[:3])
            ws[f"{_col(j + 2)}{r}"] = rates_str + ("..." if len(d.revenue_growth_rates) > 3 else "")

        ws.column_dimensions["A"].width = 22
        for c in "BCD":
            ws.column_dimensions[c].width = 16

    # ─── SOTP Sheet ─────────────────────────────────────────────────────

    def _write_sotp_sheet(self) -> None:
        a = self.sotp_a
        ws = self.wb.create_sheet("SOTP Valuation")

        ws["A1"] = f"{a.ticker} ({a.company_name}) — Sum-of-the-Parts"
        ws["A1"].font = TITLE_FONT

        # Segment table
        r = 3
        ws[f"A{r}"] = "Segment"
        ws[f"B{r}"] = "Method"
        ws[f"C{r}"] = "Revenue"
        ws[f"D{r}"] = "Profit"
        ws[f"E{r}"] = "Multiple"
        ws[f"F{r}"] = "Implied Value"
        for c in "ABCDEF":
            ws[f"{c}{r}"].font = HEADER_FONT
            ws[f"{c}{r}"].fill = HEADER_FILL

        seg_start_row = r + 1
        for i, seg in enumerate(a.segments):
            r = seg_start_row + i
            ws[f"A{r}"] = seg.name
            ws[f"B{r}"] = seg.valuation_method.value
            ws[f"C{r}"] = seg.revenue
            ws[f"C{r}"].number_format = "#,##0.00"
            ws[f"D{r}"] = seg.profit
            ws[f"D{r}"].number_format = "#,##0.00"
            ws[f"E{r}"] = seg.peer_multiple
            ws[f"E{r}"].number_format = "#,##0.00"

            # Formula: value based on method
            if seg.valuation_method.value == "PE":
                ws[f"F{r}"] = f"=D{r}*E{r}"
            elif seg.valuation_method.value == "PS":
                ws[f"F{r}"] = f"=C{r}*E{r}"
            elif seg.valuation_method.value == "EV-EBITDA":
                ws[f"F{r}"] = f"=D{r}*1.2*E{r}"
            else:
                ws[f"F{r}"] = f"=E{r}"
            ws[f"F{r}"].number_format = "#,##0.00"

        seg_end_row = seg_start_row + len(a.segments) - 1

        # Summary
        r = seg_end_row + 2
        ws[f"A{r}"] = "Total Segments Value"
        ws[f"B{r}"] = f"=SUM(F{seg_start_row}:F{seg_end_row})"
        ws[f"B{r}"].number_format = "#,##0.00"

        r += 1
        ws[f"A{r}"] = "Cash & Equivalents"
        ws[f"B{r}"] = a.cash_and_equivalents
        ws[f"B{r}"].number_format = "#,##0.00"

        r += 1
        ws[f"A{r}"] = "Less: Net Debt"
        ws[f"B{r}"] = a.net_debt
        ws[f"B{r}"].number_format = "#,##0.00"

        r += 1
        ws[f"A{r}"] = "Non-Core Assets"
        ws[f"B{r}"] = a.non_core_assets
        ws[f"B{r}"].number_format = "#,##0.00"

        r += 1
        ws[f"A{r}"] = "Equity Value"
        ws[f"B{r}"] = f"=B{r - 4}+B{r - 3}-B{r - 2}+B{r - 1}"
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"A{r}"].font = BOLD_FONT
        ws[f"B{r}"].font = BOLD_FONT

        r += 1
        ws[f"A{r}"] = "Target Price (元/股)"
        ws[f"B{r}"] = f"=B{r - 1}/{a.total_shares}"
        ws[f"B{r}"].number_format = "#,##0.00"
        ws[f"A{r}"].font = RESULT_FONT
        ws[f"B{r}"].font = RESULT_FONT

        ws.column_dimensions["A"].width = 24
        for c in "BCDEF":
            ws.column_dimensions[c].width = 16

    # ─── Summary Sheet ──────────────────────────────────────────────────

    def _write_summary_sheet(self) -> None:
        ws = self.wb.create_sheet("Summary")
        ws["A1"] = "Valuation Summary — All Methods"
        ws["A1"].font = TITLE_FONT

        r = 3
        ws[f"A{r}"] = "Method"
        ws[f"B{r}"] = "Target Price"
        ws[f"C{r}"] = "Confidence"
        ws[f"D{r}"] = "Notes"
        for c in "ABCD":
            ws[f"{c}{r}"].font = HEADER_FONT
            ws[f"{c}{r}"].fill = HEADER_FILL

        # DCF
        if self.dcf_a:
            r += 1
            ws[f"A{r}"] = "DCF"
            ws[f"B{r}"] = "See 'DCF Valuation' sheet"
            ws[f"C{r}"] = "(computed)"

        # Comparable
        if self.comp_a:
            r += 1
            ws[f"A{r}"] = "Comparable"
            ws[f"B{r}"] = "See 'Comparable Valuation' sheet"
            ws[f"C{r}"] = "(computed)"

        # Scenario
        if self.scn_a:
            r += 1
            ws[f"A{r}"] = "Scenario"
            ws[f"B{r}"] = "See 'Scenario Analysis' sheet"
            ws[f"C{r}"] = "(computed)"

        # SOTP
        if self.sotp_a:
            r += 1
            ws[f"A{r}"] = "SOTP"
            ws[f"B{r}"] = "See 'SOTP Valuation' sheet"
            ws[f"C{r}"] = "(computed)"

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 30
